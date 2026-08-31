from __future__ import annotations

import hashlib
import json
import math
import zipfile
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from capability_contracts.errors import CapabilityError, CapabilityErrorCode
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, coordinate_to_tuple
from openpyxl.utils.exceptions import InvalidFileException

from feishu_provider.sheet_delivery import (
    BorderType,
    CellStyle,
    GridRange,
    HorizontalAlignment,
    SheetDeliverySpec,
    VerticalAlignment,
)

MAX_XLSX_BYTES = 25 * 1024 * 1024
MAX_XLSX_ENTRIES = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 1_000
MAX_XLSX_TARGET_CELLS = 1_000_000


@dataclass(frozen=True, slots=True)
class XlsxVerification:
    content_hash: str
    checked_cells: int
    checked_styles: int
    checked_row_dimensions: int
    checked_column_dimensions: int
    warnings: tuple[str, ...] = ()


def verify_sheet_export(
    payload: bytes,
    *,
    target_title: str,
    spec: SheetDeliverySpec,
    retired_ranges: tuple[GridRange, ...] = (),
    neutral_row_indexes: tuple[int, ...] = (),
    neutral_column_indexes: tuple[int, ...] = (),
) -> XlsxVerification:
    _validate_archive(payload)
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=False,
            data_only=False,
            keep_links=False,
            rich_text=False,
        )
    except (InvalidFileException, KeyError, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise _verification_error("xlsx_parse_failed") from exc
    try:
        matching = [sheet for sheet in workbook.worksheets if sheet.title == target_title]
        if len(matching) != 1:
            raise _verification_error("xlsx_target_mapping_ambiguous")
        worksheet = matching[0]
        if worksheet.max_row * worksheet.max_column > MAX_XLSX_TARGET_CELLS:
            raise _verification_error("xlsx_target_grid_too_large")

        mismatches: list[str] = _worksheet_structure_mismatches(worksheet)
        actual_merges = tuple(
            sorted(
                (
                    GridRange(
                        row_start=merged.min_row - 1,
                        row_end=merged.max_row,
                        column_start=merged.min_col - 1,
                        column_end=merged.max_col,
                    )
                    for merged in worksheet.merged_cells.ranges
                ),
                key=_range_key,
            )
        )
        expected_merges = tuple(sorted(spec.merges, key=_range_key))
        if actual_merges != expected_merges:
            mismatches.append("merge_ranges")

        frozen_rows, frozen_columns = _freeze_counts(worksheet.freeze_panes)
        if frozen_rows != spec.frozen_row_count:
            mismatches.append("frozen_rows")
        if frozen_columns != spec.frozen_column_count:
            mismatches.append("frozen_columns")

        expected_values = spec.remote_values()
        for row_index in range(spec.row_count):
            for column_index in range(spec.column_count):
                actual = worksheet.cell(row_index + 1, column_index + 1).value
                expected = expected_values[row_index][column_index]
                if not _values_equal(actual, expected):
                    mismatches.append(
                        f"cell_value:{row_index}:{column_index}"
                    )
                cell = worksheet.cell(row_index + 1, column_index + 1)
                if cell.hyperlink is not None:
                    mismatches.append(
                        f"cell_hyperlink:{row_index}:{column_index}"
                    )
                if cell.comment is not None:
                    mismatches.append(
                        f"cell_comment:{row_index}:{column_index}"
                    )

        for row in worksheet.iter_rows():
            for cell in row:
                if (
                    cell.row <= spec.row_count
                    and cell.column <= spec.column_count
                ):
                    continue
                if (
                    not _empty_value(cell.value)
                    or cell.hyperlink is not None
                    or cell.comment is not None
                ):
                    mismatches.append("content_outside_delivery_rectangle")
                    break
            if "content_outside_delivery_rectangle" in mismatches:
                break

        merge_by_cell = _merge_index(expected_merges)
        for row_index in range(spec.row_count):
            for column_index in range(spec.column_count):
                cell = worksheet.cell(row_index + 1, column_index + 1)
                merge = merge_by_cell.get((row_index, column_index))
                style, style_range = _expected_style_context(
                    spec, row_index, column_index
                )
                if merge is None or (
                    row_index == merge.row_start
                    and column_index == merge.column_start
                ):
                    mismatches.extend(
                        _style_mismatches(
                            cell,
                            style,
                            row_index=row_index,
                            column_index=column_index,
                            style_range=style_range,
                            compare_border=merge is None,
                        )
                    )
                if merge is not None:
                    mismatches.extend(
                        _merged_border_mismatches(
                            cell,
                            spec,
                            merge=merge,
                            row_index=row_index,
                            column_index=column_index,
                        )
                    )

        row_sizes = _dimension_sizes(
            spec.row_count,
            spec.default_row_height_px,
            spec.row_heights,
        )
        for row_number, expected_pixels in enumerate(row_sizes, start=1):
            dimension = worksheet.row_dimensions.get(row_number)
            points = (
                dimension.height
                if dimension is not None and dimension.height is not None
                else worksheet.sheet_format.defaultRowHeight
            )
            actual_pixels = None if points is None else _points_to_pixels(points)
            if (
                actual_pixels is None
                or abs(actual_pixels - expected_pixels) > 2
            ):
                mismatches.append(
                    f"row_height:{row_number - 1}:"
                    f"actual_{actual_pixels if actual_pixels is not None else 'none'}:"
                    f"expected_{expected_pixels}"
                )
            if dimension is not None and bool(dimension.hidden):
                mismatches.append(f"row_hidden:{row_number - 1}")

        column_sizes = _dimension_sizes(
            spec.column_count,
            spec.default_column_width_px,
            spec.column_widths,
        )
        max_digit_width = _column_max_digit_width(workbook)
        if max_digit_width is None:
            mismatches.append("column_width_font_unsupported")
        for column_number, expected_pixels in enumerate(column_sizes, start=1):
            width = _column_width(worksheet, column_number)
            actual_pixels = (
                None
                if width is None or max_digit_width is None
                else _column_width_to_pixels(width, max_digit_width)
            )
            if (
                actual_pixels is None
                or abs(actual_pixels - expected_pixels)
                > _column_width_tolerance(max_digit_width)
            ):
                mismatches.append(
                    f"column_width:{column_number - 1}:"
                    f"raw_milli_{round(width * 1000) if width is not None else 'none'}:"
                    f"actual_{actual_pixels if actual_pixels is not None else 'none'}:"
                    f"expected_{expected_pixels}"
                )
            dimension = _column_dimension(worksheet, column_number)
            if dimension is not None and bool(dimension.hidden):
                mismatches.append(f"column_hidden:{column_number - 1}")

        for retired in retired_ranges:
            if retired.overlaps(spec.delivery_range):
                mismatches.append("retired_range_overlaps_delivery_rectangle")
                continue
            for row_index in range(retired.row_start, retired.row_end):
                for column_index in range(
                    retired.column_start, retired.column_end
                ):
                    cell = worksheet.cell(row_index + 1, column_index + 1)
                    if (
                        not _empty_value(cell.value)
                        or cell.hyperlink is not None
                        or cell.comment is not None
                    ):
                        mismatches.append(
                            f"retired_content:{row_index}:{column_index}"
                        )
                    mismatches.extend(
                        _retired_style_mismatches(
                            cell,
                            row_index=row_index,
                            column_index=column_index,
                        )
                    )

        for row_index in neutral_row_indexes:
            if row_index < spec.row_count:
                mismatches.append(f"neutral_row_inside_delivery:{row_index}")
                continue
            dimension = worksheet.row_dimensions.get(row_index + 1)
            points = (
                dimension.height
                if dimension is not None and dimension.height is not None
                else worksheet.sheet_format.defaultRowHeight
            )
            actual_pixels = None if points is None else _points_to_pixels(points)
            if actual_pixels is None or abs(actual_pixels - 24) > 2:
                mismatches.append(
                    f"retired_row_height:{row_index}:"
                    f"actual_{actual_pixels if actual_pixels is not None else 'none'}:"
                    "expected_24"
                )
            if dimension is not None and bool(dimension.hidden):
                mismatches.append(f"retired_row_hidden:{row_index}")

        for column_index in neutral_column_indexes:
            if column_index < spec.column_count:
                mismatches.append(
                    f"neutral_column_inside_delivery:{column_index}"
                )
                continue
            width = _column_width(worksheet, column_index + 1)
            actual_pixels = (
                None
                if width is None or max_digit_width is None
                else _column_width_to_pixels(width, max_digit_width)
            )
            if (
                actual_pixels is None
                or abs(actual_pixels - 100)
                > _column_width_tolerance(max_digit_width)
            ):
                mismatches.append(
                    f"retired_column_width:{column_index}:"
                    f"actual_{actual_pixels if actual_pixels is not None else 'none'}:"
                    "expected_100"
                )
            dimension = _column_dimension(worksheet, column_index + 1)
            if dimension is not None and bool(dimension.hidden):
                mismatches.append(f"retired_column_hidden:{column_index}")

        if mismatches:
            raise _verification_error(*_deduplicate(mismatches)[:50])

        canonical = json.dumps(
            {
                "target_title": target_title,
                "spec_hash": spec.content_hash,
                "rows": spec.row_count,
                "columns": spec.column_count,
                "merges": [item.model_dump(mode="json") for item in actual_merges],
                "frozen_rows": frozen_rows,
                "frozen_columns": frozen_columns,
                "retired_ranges": [
                    item.model_dump(mode="json") for item in retired_ranges
                ],
                "neutral_rows": list(neutral_row_indexes),
                "neutral_columns": list(neutral_column_indexes),
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        return XlsxVerification(
            content_hash="sha256:" + hashlib.sha256(canonical).hexdigest(),
            checked_cells=spec.row_count * spec.column_count,
            checked_styles=(
                spec.row_count * spec.column_count
                + sum(item.cell_count for item in retired_ranges)
            ),
            checked_row_dimensions=(
                spec.row_count + len(neutral_row_indexes)
            ),
            checked_column_dimensions=(
                spec.column_count + len(neutral_column_indexes)
            ),
        )
    finally:
        workbook.close()


def _validate_archive(payload: bytes) -> None:
    if not payload or len(payload) > MAX_XLSX_BYTES:
        raise _verification_error("xlsx_size_limit")
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ENTRIES:
                raise _verification_error("xlsx_entry_limit")
            total_uncompressed = 0
            for entry in entries:
                total_uncompressed += entry.file_size
                if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise _verification_error("xlsx_uncompressed_size_limit")
                if (
                    entry.file_size > 0
                    and entry.compress_size == 0
                ):
                    raise _verification_error("xlsx_invalid_compression")
                if (
                    entry.compress_size > 0
                    and entry.file_size / entry.compress_size
                    > MAX_XLSX_COMPRESSION_RATIO
                ):
                    raise _verification_error("xlsx_compression_ratio_limit")
                normalized = entry.filename.replace("\\", "/")
                if normalized.startswith("/") or ".." in normalized.split("/"):
                    raise _verification_error("xlsx_unsafe_entry_name")
            if "xl/workbook.xml" not in {entry.filename for entry in entries}:
                raise _verification_error("xlsx_workbook_missing")
    except zipfile.BadZipFile as exc:
        raise _verification_error("xlsx_archive_invalid") from exc


def _expected_style_context(
    spec: SheetDeliverySpec,
    row_index: int,
    column_index: int,
) -> tuple[CellStyle, GridRange]:
    for range_, style in reversed(spec.resolved_style_ranges()):
        if range_.contains(row_index, column_index):
            return style, range_
    return spec.base_style, spec.delivery_range


def _style_mismatches(
    cell: Any,
    expected: CellStyle,
    *,
    row_index: int,
    column_index: int,
    style_range: GridRange,
    compare_border: bool,
) -> list[str]:
    prefix = f"cell_style:{row_index}:{column_index}"
    mismatches: list[str] = []
    if bool(cell.font.bold) != expected.bold:
        mismatches.append(prefix + ":bold")
    if bool(cell.font.italic) != expected.italic:
        mismatches.append(prefix + ":italic")
    if cell.font.sz is None or not math.isclose(
        float(cell.font.sz), expected.font_size_pt, abs_tol=0.25
    ):
        mismatches.append(prefix + ":font_size")
    if bool(cell.font.underline) != expected.underline:
        mismatches.append(prefix + ":underline")
    if bool(cell.font.strike) != expected.strikethrough:
        mismatches.append(prefix + ":strikethrough")
    if _color(cell.font.color) != expected.text_color:
        mismatches.append(prefix + ":text_color")
    actual_fill_color = _fill_color(cell.fill)
    if actual_fill_color != expected.fill_color and not (
        actual_fill_color is None and expected.fill_color == "#FFFFFF"
    ):
        mismatches.append(prefix + ":fill_color")
    expected_horizontal = {
        HorizontalAlignment.LEFT: "left",
        HorizontalAlignment.CENTER: "center",
        HorizontalAlignment.RIGHT: "right",
    }[expected.horizontal_alignment]
    expected_vertical = {
        VerticalAlignment.TOP: "top",
        VerticalAlignment.MIDDLE: "center",
        VerticalAlignment.BOTTOM: "bottom",
    }[expected.vertical_alignment]
    if (cell.alignment.horizontal or "left") != expected_horizontal:
        mismatches.append(prefix + ":horizontal_alignment")
    if (cell.alignment.vertical or "top") != expected_vertical:
        mismatches.append(prefix + ":vertical_alignment")
    if bool(cell.alignment.wrap_text) != expected.wrap_text:
        mismatches.append(prefix + ":wrap_text")
    if cell.number_format != "General":
        mismatches.append(prefix + ":number_format")
    if compare_border:
        expected_edges = _expected_edges(
            expected.border_type,
            style_range,
            row_index,
            column_index,
        )
        mismatches.extend(
            _border_mismatches(cell, expected_edges, expected.border_color, prefix)
        )
    return mismatches


def _retired_style_mismatches(
    cell: Any,
    *,
    row_index: int,
    column_index: int,
) -> list[str]:
    prefix = f"retired_style:{row_index}:{column_index}"
    mismatches: list[str] = []
    if bool(cell.font.bold):
        mismatches.append(prefix + ":bold")
    if bool(cell.font.italic):
        mismatches.append(prefix + ":italic")
    if bool(cell.font.underline):
        mismatches.append(prefix + ":underline")
    if bool(cell.font.strike):
        mismatches.append(prefix + ":strikethrough")
    if _fill_color(cell.fill) is not None:
        mismatches.append(prefix + ":fill_color")
    if any(
        getattr(cell.border, edge) is not None
        and getattr(cell.border, edge).style is not None
        for edge in ("left", "right", "top", "bottom")
    ):
        mismatches.append(prefix + ":border")
    if cell.alignment.horizontal not in {None, "general", "left"}:
        mismatches.append(prefix + ":horizontal_alignment")
    if cell.alignment.vertical not in {None, "top"}:
        mismatches.append(prefix + ":vertical_alignment")
    if bool(cell.alignment.wrap_text):
        mismatches.append(prefix + ":wrap_text")
    if cell.number_format != "General":
        mismatches.append(prefix + ":number_format")
    return mismatches


def _merged_border_mismatches(
    cell: Any,
    spec: SheetDeliverySpec,
    *,
    merge: GridRange,
    row_index: int,
    column_index: int,
) -> list[str]:
    anchor_style, anchor_range = _expected_style_context(
        spec,
        merge.row_start,
        merge.column_start,
    )
    end_row = merge.row_end - 1
    end_column = merge.column_end - 1
    end_style, end_range = _expected_style_context(spec, end_row, end_column)
    anchor_edges = _expected_edges(
        anchor_style.border_type,
        anchor_range,
        merge.row_start,
        merge.column_start,
    )
    end_edges = _expected_edges(
        end_style.border_type,
        end_range,
        end_row,
        end_column,
    )
    outer_edges = {
        "left": anchor_edges["left"],
        "right": end_edges["right"],
        "top": anchor_edges["top"],
        "bottom": end_edges["bottom"],
    }
    outer_colors = {
        "left": anchor_style.border_color,
        "right": end_style.border_color,
        "top": anchor_style.border_color,
        "bottom": end_style.border_color,
    }
    is_anchor = (
        row_index == merge.row_start and column_index == merge.column_start
    )
    visible_edges = (
        outer_edges
        if is_anchor
        else {
            "left": outer_edges["left"] and column_index == merge.column_start,
            "right": outer_edges["right"] and column_index == end_column,
            "top": outer_edges["top"] and row_index == merge.row_start,
            "bottom": outer_edges["bottom"] and row_index == end_row,
        }
    )
    return _border_mismatches(
        cell,
        visible_edges,
        outer_colors,
        f"merged_border:{row_index}:{column_index}",
    )


def _expected_edges(
    border_type: BorderType,
    range_: GridRange,
    row_index: int,
    column_index: int,
) -> dict[str, bool]:
    edges = {"left": False, "right": False, "top": False, "bottom": False}
    if border_type is BorderType.FULL:
        return {edge: True for edge in edges}
    if border_type is BorderType.OUTER:
        edges.update(
            left=column_index == range_.column_start,
            right=column_index == range_.column_end - 1,
            top=row_index == range_.row_start,
            bottom=row_index == range_.row_end - 1,
        )
    elif border_type is BorderType.INNER:
        edges.update(
            left=column_index > range_.column_start,
            right=column_index < range_.column_end - 1,
            top=row_index > range_.row_start,
            bottom=row_index < range_.row_end - 1,
        )
    elif border_type.value in edges:
        edges[border_type.value] = True
    return edges


def _border_mismatches(
    cell: Any,
    expected_edges: dict[str, bool],
    expected_color: str | None | dict[str, str | None],
    prefix: str,
) -> list[str]:
    mismatches: list[str] = []
    for edge, expected_present in expected_edges.items():
        side = getattr(cell.border, edge)
        actual_present = side is not None and side.style is not None
        if actual_present != expected_present:
            mismatches.append(f"{prefix}:border_{edge}")
        elif expected_present and _color(side.color) != (
            expected_color[edge]
            if isinstance(expected_color, dict)
            else expected_color
        ):
            mismatches.append(f"{prefix}:border_{edge}_color")
    return mismatches


def _color(value: Any) -> str | None:
    if value is None:
        return None
    if getattr(value, "type", None) == "rgb" and value.rgb:
        raw = str(value.rgb).upper()
        return "#" + raw[-6:]
    return None


def _fill_color(fill: Any) -> str | None:
    if fill is None or fill.fill_type is None:
        return None
    if fill.fill_type != "solid":
        return None
    return _color(fill.fgColor)


def _freeze_counts(value: Any) -> tuple[int, int]:
    if value is None:
        return 0, 0
    coordinate = value.coordinate if hasattr(value, "coordinate") else str(value)
    row, column = coordinate_to_tuple(coordinate)
    return row - 1, column - 1


def _merge_index(
    merges: tuple[GridRange, ...],
) -> dict[tuple[int, int], GridRange]:
    result: dict[tuple[int, int], GridRange] = {}
    for merge in merges:
        for row_index in range(merge.row_start, merge.row_end):
            for column_index in range(merge.column_start, merge.column_end):
                result[(row_index, column_index)] = merge
    return result


def _dimension_sizes(extent: int, default: int, spans: tuple[Any, ...]) -> list[int]:
    result = [default] * extent
    for span in spans:
        for index in range(span.start_index, span.end_index):
            result[index] = span.pixel_size
    return result


def _points_to_pixels(points: float) -> int:
    return round(float(points) * 96 / 72)


def _column_width(worksheet: Any, column_number: int) -> float | None:
    dimension = _column_dimension(worksheet, column_number)
    if dimension is not None and dimension.width is not None:
        return float(dimension.width)
    default_width = worksheet.sheet_format.defaultColWidth
    if default_width is not None:
        return float(default_width)
    base_width = worksheet.sheet_format.baseColWidth
    return float(base_width) if base_width is not None else None


def _column_dimension(worksheet: Any, column_number: int) -> Any | None:
    for dimension in worksheet.column_dimensions.values():
        start = dimension.min or column_index_from_string(dimension.index)
        end = dimension.max or start
        if start <= column_number <= end:
            return dimension
    return None


def _worksheet_structure_mismatches(worksheet: Any) -> list[str]:
    mismatches: list[str] = []
    if getattr(worksheet, "_charts", ()):
        mismatches.append("worksheet_charts")
    if getattr(worksheet, "_images", ()):
        mismatches.append("worksheet_images")
    if len(worksheet.tables):
        mismatches.append("worksheet_tables")
    if worksheet.auto_filter.ref:
        mismatches.append("worksheet_auto_filter")
    if len(worksheet.conditional_formatting):
        mismatches.append("worksheet_conditional_formatting")
    if worksheet.data_validations.dataValidation:
        mismatches.append("worksheet_data_validations")
    if bool(worksheet.protection.sheet):
        mismatches.append("worksheet_protection")
    if worksheet.sheet_state != "visible":
        mismatches.append("worksheet_not_visible")
    return mismatches


def _column_max_digit_width(workbook: Any) -> float | None:
    try:
        normal = workbook._named_styles["Normal"]
        font_name = str(normal.font.name or "").casefold()
        font_size = float(normal.font.sz)
    except (AttributeError, KeyError, TypeError, ValueError):
        return None
    if font_name != "calibri" or not math.isfinite(font_size):
        return None
    if font_size < 1 or font_size > 100:
        return None
    return 7.0 * font_size / 11.0


def _column_width_tolerance(max_digit_width: float | None) -> int:
    if max_digit_width is None:
        return 0
    # Feishu quantizes fixed pixel widths to whole OOXML character-width units
    # during export. One unit is one maximum digit width in the Normal font.
    return max(2, math.ceil(max_digit_width))


def _column_width_to_pixels(width: float, max_digit_width: float) -> int:
    if width < 1:
        return round(width * (max_digit_width + 5))
    return round(width * max_digit_width + 5)


def _values_equal(actual: Any, expected: Any) -> bool:
    if _empty_value(actual) and _empty_value(expected):
        return True
    if isinstance(actual, bool) or isinstance(expected, bool):
        return actual is expected
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return math.isclose(float(actual), float(expected), rel_tol=0, abs_tol=1e-12)
    return actual == expected


def _empty_value(value: Any) -> bool:
    return value is None or value == ""


def _range_key(value: GridRange) -> tuple[int, int, int, int]:
    return (
        value.row_start,
        value.column_start,
        value.row_end,
        value.column_end,
    )


def _deduplicate(values: list[str]) -> list[str]:
    return list(dict.fromkeys(values))


def _verification_error(*mismatches: str) -> CapabilityError:
    return CapabilityError(
        CapabilityErrorCode.VERIFICATION_INCOMPLETE,
        "The exported XLSX does not prove the authorized worksheet state.",
        details={"mismatches": list(mismatches)},
    )
