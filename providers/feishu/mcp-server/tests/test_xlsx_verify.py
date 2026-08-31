from __future__ import annotations

import zipfile
from io import BytesIO

import pytest
from capability_contracts import CapabilityError, CapabilityErrorCode
from feishu_provider.sheet_delivery import (
    SHEET_DELIVERY_SCHEMA_VERSION,
    BorderType,
    GridRange,
    SheetDeliverySpec,
)
from feishu_provider.xlsx_verify import verify_sheet_export
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _spec(**overrides: object) -> SheetDeliverySpec:
    payload: dict[str, object] = {
        "schema_version": SHEET_DELIVERY_SCHEMA_VERSION,
        "row_count": 2,
        "column_count": 2,
        "values": [["标题", "值"], [1, {"formula": "=A2+1"}]],
        "base_style": {
            "font_size_pt": 10,
            "text_color": "#000000",
            "border_type": "full",
            "border_color": "#D0D5DD",
            "horizontal_alignment": "left",
            "vertical_alignment": "middle",
            "wrap_text": False,
        },
        "style_ranges": [
            {
                "range": {
                    "row_start": 0,
                    "row_end": 1,
                    "column_start": 0,
                    "column_end": 2,
                },
                "style": {"bold": True, "fill_color": "#F2F4F7"},
            }
        ],
        "merges": [],
        "default_row_height_px": 24,
        "default_column_width_px": 100,
        "row_heights": [
            {"start_index": 0, "end_index": 1, "pixel_size": 32}
        ],
        "column_widths": [],
        "frozen_row_count": 1,
        "frozen_column_count": 0,
    }
    payload.update(overrides)
    return SheetDeliverySpec.model_validate(payload)


def _argb(color: str | None) -> str | None:
    return None if color is None else "FF" + color.removeprefix("#")


def _border(style: object) -> Border:
    border_type = style.border_type
    color = _argb(style.border_color)
    present = Side(style="thin", color=color)
    absent = Side()
    if border_type is BorderType.FULL:
        return Border(left=present, right=present, top=present, bottom=present)
    return Border(left=absent, right=absent, top=absent, bottom=absent)


def _workbook_bytes(
    spec: SheetDeliverySpec,
    *,
    outside_value: object | None = None,
    inside_hyperlink: bool = False,
    add_chart: bool = False,
    normal_font_size_pt: float = 11,
    raw_column_widths: tuple[float, ...] | None = None,
) -> bytes:
    workbook = Workbook()
    workbook._named_styles["Normal"].font = Font(
        name="Calibri",
        size=normal_font_size_pt,
    )
    worksheet = workbook.active
    worksheet.title = "交付"
    worksheet.freeze_panes = (
        worksheet.cell(spec.frozen_row_count + 1, spec.frozen_column_count + 1)
        if spec.frozen_row_count or spec.frozen_column_count
        else None
    )
    values = spec.remote_values()
    for row_index in range(spec.row_count):
        for column_index in range(spec.column_count):
            cell = worksheet.cell(row_index + 1, column_index + 1)
            cell.value = values[row_index][column_index]
            style = spec.style_at(row_index, column_index)
            cell.font = Font(
                bold=style.bold,
                italic=style.italic,
                size=style.font_size_pt,
                underline="single" if style.underline else None,
                strike=style.strikethrough,
                color=_argb(style.text_color),
            )
            cell.fill = (
                PatternFill(fill_type="solid", fgColor=_argb(style.fill_color))
                if style.fill_color
                else PatternFill()
            )
            cell.alignment = Alignment(
                horizontal=style.horizontal_alignment.value,
                vertical=(
                    "center"
                    if style.vertical_alignment.value == "middle"
                    else style.vertical_alignment.value
                ),
                wrap_text=style.wrap_text,
            )
            cell.border = _border(style)
    for merge in spec.merges:
        worksheet.merge_cells(
            start_row=merge.row_start + 1,
            end_row=merge.row_end,
            start_column=merge.column_start + 1,
            end_column=merge.column_end,
        )
    row_sizes = [spec.default_row_height_px] * spec.row_count
    for span in spec.row_heights:
        row_sizes[span.start_index : span.end_index] = [span.pixel_size] * (
            span.end_index - span.start_index
        )
    for row_number, pixels in enumerate(row_sizes, start=1):
        worksheet.row_dimensions[row_number].height = pixels * 72 / 96
    column_sizes = [spec.default_column_width_px] * spec.column_count
    for span in spec.column_widths:
        column_sizes[span.start_index : span.end_index] = [span.pixel_size] * (
            span.end_index - span.start_index
        )
    for column_number, pixels in enumerate(column_sizes, start=1):
        worksheet.column_dimensions[
            chr(ord("A") + column_number - 1)
        ].width = (
            raw_column_widths[column_number - 1]
            if raw_column_widths is not None
            else (pixels - 5) / 7 if pixels >= 12 else pixels / 12
        )
    if outside_value is not None:
        worksheet.cell(spec.row_count + 1, 1).value = outside_value
    if inside_hyperlink:
        worksheet.cell(1, 1).hyperlink = "https://example.com"
    if add_chart:
        chart = BarChart()
        chart.add_data(Reference(worksheet, min_col=1, min_row=2, max_row=2))
        worksheet.add_chart(chart, "D1")
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _revision_workbook_bytes(
    spec: SheetDeliverySpec,
    *,
    dirty_retired_style: bool = False,
    explicit_neutral_alignment: bool = False,
) -> bytes:
    workbook = load_workbook(BytesIO(_workbook_bytes(spec)))
    worksheet = workbook["交付"]
    worksheet.row_dimensions[3].height = 18
    worksheet.column_dimensions["C"].width = (100 - 5) / 7
    if dirty_retired_style:
        worksheet["C1"].font = Font(bold=True)
    if explicit_neutral_alignment:
        for coordinate in ("C1", "C2", "A3", "B3", "C3"):
            worksheet[coordinate].alignment = Alignment(
                horizontal="left",
                vertical="top",
            )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _without_fills(payload: bytes, *coordinates: str) -> bytes:
    workbook = load_workbook(BytesIO(payload))
    worksheet = workbook["交付"]
    for coordinate in coordinates:
        worksheet[coordinate].fill = PatternFill()
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_export_verification_proves_values_styles_dimensions_and_freeze() -> None:
    spec = _spec()

    result = verify_sheet_export(
        _workbook_bytes(spec),
        target_title="交付",
        spec=spec,
    )

    assert result.content_hash.startswith("sha256:")
    assert result.checked_cells == 4
    assert result.checked_styles == 4
    assert result.checked_row_dimensions == 2
    assert result.checked_column_dimensions == 2


def test_export_verification_uses_normal_font_for_column_width_units() -> None:
    spec = _spec(
        column_count=3,
        values=[["标题", "值", "说明"], [1, 2, 3]],
        column_widths=[
            {"start_index": 0, "end_index": 1, "pixel_size": 150},
            {"start_index": 1, "end_index": 2, "pixel_size": 100},
            {"start_index": 2, "end_index": 3, "pixel_size": 300},
        ],
    )

    result = verify_sheet_export(
        _workbook_bytes(
            spec,
            normal_font_size_pt=12,
            raw_column_widths=(19, 12, 39),
        ),
        target_title="交付",
        spec=spec,
    )

    assert result.checked_column_dimensions == 3


def test_export_verification_accepts_one_feishu_character_width_quantum() -> None:
    spec = _spec(
        column_count=1,
        values=[["标题"], [1]],
        style_ranges=[],
        column_widths=[
            {"start_index": 0, "end_index": 1, "pixel_size": 72},
        ],
    )

    result = verify_sheet_export(
        _workbook_bytes(
            spec,
            normal_font_size_pt=12,
            raw_column_widths=(8,),
        ),
        target_title="交付",
        spec=spec,
    )

    assert result.checked_column_dimensions == 1


def test_export_verification_accepts_white_fill_normalized_to_none() -> None:
    base_style = _spec().base_style.model_dump(mode="json")
    base_style["fill_color"] = "#FFFFFF"
    spec = _spec(base_style=base_style, style_ranges=[])

    result = verify_sheet_export(
        _without_fills(_workbook_bytes(spec), "A1", "B1", "A2", "B2"),
        target_title="交付",
        spec=spec,
    )

    assert result.checked_styles == 4


def test_export_verification_rejects_nonwhite_fill_normalized_to_none() -> None:
    spec = _spec()

    with pytest.raises(CapabilityError) as error:
        verify_sheet_export(
            _without_fills(_workbook_bytes(spec), "A1"),
            target_title="交付",
            spec=spec,
        )

    assert "cell_style:0:0:fill_color" in error.value.details["mismatches"]


def test_export_verification_still_rejects_materially_wrong_column_width() -> None:
    spec = _spec(
        column_widths=[
            {"start_index": 0, "end_index": 1, "pixel_size": 150},
        ],
    )

    with pytest.raises(CapabilityError) as error:
        verify_sheet_export(
            _workbook_bytes(
                spec,
                normal_font_size_pt=12,
                raw_column_widths=(12, 12),
            ),
            target_title="交付",
            spec=spec,
        )

    assert any(
        mismatch.startswith("column_width:0:")
        for mismatch in error.value.details["mismatches"]
    )


def test_export_verification_supports_a_merge_with_blank_subordinate_cells() -> None:
    spec = _spec(
        values=[["合并标题", None], [1, 2]],
        merges=[
            {
                "row_start": 0,
                "row_end": 1,
                "column_start": 0,
                "column_end": 2,
            }
        ],
        base_style={"font_size_pt": 10, "text_color": "#000000"},
        style_ranges=[],
    )

    result = verify_sheet_export(
        _workbook_bytes(spec),
        target_title="交付",
        spec=spec,
    )

    assert result.checked_cells == 4


def test_export_verification_accepts_anchor_aggregated_merged_border() -> None:
    spec = _spec(
        values=[["合并标题", None], [1, 2]],
        merges=[
            {
                "row_start": 0,
                "row_end": 1,
                "column_start": 0,
                "column_end": 2,
            }
        ],
    )

    result = verify_sheet_export(
        _workbook_bytes(spec),
        target_title="交付",
        spec=spec,
    )

    assert result.checked_cells == 4


def test_export_verification_rejects_content_outside_delivery_rectangle() -> None:
    spec = _spec()

    with pytest.raises(CapabilityError) as error:
        verify_sheet_export(
            _workbook_bytes(spec, outside_value="残留内容"),
            target_title="交付",
            spec=spec,
        )

    assert error.value.code is CapabilityErrorCode.VERIFICATION_INCOMPLETE
    assert "content_outside_delivery_rectangle" in error.value.details["mismatches"]


def test_revision_export_verifies_retired_styles_and_neutral_dimensions() -> None:
    spec = _spec()
    retired = (
        GridRange(row_start=2, row_end=3, column_start=0, column_end=3),
        GridRange(row_start=0, row_end=2, column_start=2, column_end=3),
    )
    result = verify_sheet_export(
        _revision_workbook_bytes(spec),
        target_title="交付",
        spec=spec,
        retired_ranges=retired,
        neutral_row_indexes=(2,),
        neutral_column_indexes=(2,),
    )

    assert result.content_hash.startswith("sha256:")


def test_revision_export_accepts_explicit_platform_neutral_alignment() -> None:
    spec = _spec()
    retired = (
        GridRange(row_start=2, row_end=3, column_start=0, column_end=3),
        GridRange(row_start=0, row_end=2, column_start=2, column_end=3),
    )

    result = verify_sheet_export(
        _revision_workbook_bytes(spec, explicit_neutral_alignment=True),
        target_title="交付",
        spec=spec,
        retired_ranges=retired,
        neutral_row_indexes=(2,),
        neutral_column_indexes=(2,),
    )

    assert result.content_hash.startswith("sha256:")


def test_revision_export_rejects_uncleared_retired_style() -> None:
    spec = _spec()
    with pytest.raises(CapabilityError) as error:
        verify_sheet_export(
            _revision_workbook_bytes(spec, dirty_retired_style=True),
            target_title="交付",
            spec=spec,
            retired_ranges=(
                GridRange(
                    row_start=0,
                    row_end=2,
                    column_start=2,
                    column_end=3,
                ),
            ),
            neutral_column_indexes=(2,),
        )

    assert "retired_style:0:2:bold" in error.value.details["mismatches"]


@pytest.mark.parametrize(
    ("options", "mismatch"),
    [
        ({"inside_hyperlink": True}, "cell_hyperlink:0:0"),
        ({"add_chart": True}, "worksheet_charts"),
    ],
)
def test_export_verification_rejects_undeclared_business_structures(
    options: dict[str, bool],
    mismatch: str,
) -> None:
    spec = _spec()

    with pytest.raises(CapabilityError) as error:
        verify_sheet_export(
            _workbook_bytes(spec, **options),
            target_title="交付",
            spec=spec,
        )

    assert mismatch in error.value.details["mismatches"]


def test_export_verification_rejects_an_unsafe_archive() -> None:
    payload = BytesIO()
    with zipfile.ZipFile(payload, mode="w") as archive:
        archive.writestr("../escape", "unsafe")
        archive.writestr("xl/workbook.xml", "<workbook/>")

    with pytest.raises(CapabilityError) as error:
        verify_sheet_export(payload.getvalue(), target_title="交付", spec=_spec())

    assert error.value.details == {"mismatches": ["xlsx_unsafe_entry_name"]}
