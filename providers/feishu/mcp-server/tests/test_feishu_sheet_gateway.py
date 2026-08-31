from __future__ import annotations

import asyncio
import json
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from io import BytesIO

import httpx
import pytest
from capability_contracts import CapabilityError, CapabilityErrorCode
from feishu_protocol import (
    SHEETS_EXPORT_VERIFY_CAPABILITY,
    SHEETS_MANAGED_WRITE_CAPABILITY,
    SHEETS_READ_CAPABILITY,
    SHEETS_TYPED_VALUES_WRITE_CAPABILITY,
    WIKI_CHILD_LIST_CAPABILITY,
    WIKI_NODE_CREATE_CAPABILITY,
    WIKI_NODE_READ_CAPABILITY,
)
from feishu_provider.feishu_sheet_gateway import FeishuManagedSheetsGateway
from feishu_provider.lease_client import ProviderTokenLease
from feishu_provider.managed_sheets import RemoteMutationFailure
from feishu_provider.operation_store import (
    ManagedSheetRegistration,
    OperationRecord,
    OperationState,
    ProtectedTarget,
    RevisionRecord,
    RevisionState,
    RevisionStep,
    WriteStep,
)
from feishu_provider.sheet_delivery import (
    SHEET_DELIVERY_SCHEMA_VERSION,
    PlacementMode,
    SheetDeliverySpec,
)
from feishu_provider.sheet_revision import build_revision_diff
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

SHEET_TOKEN = "shtcn1234567890"
SHEET_ID = "sheet-one"
PROFILE_REF = "profile_0123456789abcdef0123"
LOCATOR = f"https://example.feishu.cn/sheets/{SHEET_TOKEN}?sheet={SHEET_ID}"
WIKI_SPACE_ID = "7527507018619224092"
WIKI_PARENT_TOKEN = "MlK5wn103ikcd1kA1JScXTFCnOb"
WIKI_PARENT_LOCATOR = f"https://example.feishu.cn/wiki/{WIKI_PARENT_TOKEN}"
WIKI_PARENT_OBJECT_TOKEN = "doxcnParentObject123"
CREATED_WIKI_NODE_TOKEN = "wikcnCreatedWorkbook123"
CREATED_SHEET_TOKEN = "shtcnCreatedWorkbook123"


class FakeLeaseClient:
    def __init__(self) -> None:
        self.calls: list[tuple[str, str | None, tuple[str, ...]]] = []

    async def issue(
        self,
        *,
        task_ref: str,
        profile_ref: str | None,
        capabilities: tuple[str, ...],
    ) -> ProviderTokenLease:
        self.calls.append((task_ref, profile_ref, capabilities))
        return ProviderTokenLease(
            lease_ref="lease_write_test",
            task_ref=task_ref,
            profile_ref=profile_ref or PROFILE_REF,
            capabilities=capabilities,
            scopes=(
                "sheets:spreadsheet",
                "sheets:spreadsheet:write_only",
                "drive:export:readonly",
            ),
            access_token="write-access-token-must-not-render",
            issued_at="2026-08-25T00:00:00+00:00",
            expires_at="2026-08-25T00:10:00+00:00",
            token_expires_at="2026-08-25T01:00:00+00:00",
        )

    async def aclose(self) -> None:
        return None


def _spec() -> SheetDeliverySpec:
    return SheetDeliverySpec.model_validate(
        {
            "schema_version": SHEET_DELIVERY_SCHEMA_VERSION,
            "row_count": 2,
            "column_count": 2,
            "values": [["标题", "值"], [1, 2]],
            "base_style": {
                "font_size_pt": 10,
                "text_color": "#000000",
                "horizontal_alignment": "left",
                "vertical_alignment": "top",
            },
            "default_row_height_px": 24,
            "default_column_width_px": 100,
        }
    )


def _boolean_spec() -> SheetDeliverySpec:
    return SheetDeliverySpec.model_validate(
        {
            **_spec().model_dump(mode="json"),
            "values": [["布尔", None], [True, 2]],
        }
    )


def _metadata() -> dict[str, object]:
    return {
        "code": 0,
        "data": {
            "spreadsheet": {
                "title": "目标工作簿",
                "token": SHEET_TOKEN,
                "url": f"https://example.feishu.cn/sheets/{SHEET_TOKEN}",
            }
        },
    }


def _sheet(*, values_written: bool = False) -> dict[str, object]:
    del values_written
    return {
        "sheet_id": SHEET_ID,
        "title": "空白页",
        "index": 0,
        "hidden": False,
        "resource_type": "sheet",
        "grid_properties": {
            "row_count": 2,
            "column_count": 2,
            "frozen_row_count": 0,
            "frozen_column_count": 0,
        },
        "merges": [],
    }


def _wiki_node(
    *,
    node_token: str,
    object_token: str,
    object_type: str,
    title: str,
    parent_node_token: str | None = WIKI_PARENT_TOKEN,
) -> dict[str, object]:
    return {
        "space_id": WIKI_SPACE_ID,
        "node_token": node_token,
        "obj_token": object_token,
        "obj_type": object_type,
        "parent_node_token": parent_node_token or "",
        "node_type": "origin",
        "has_child": False,
        "title": title,
    }


def _wiki_parent() -> dict[str, object]:
    return _wiki_node(
        node_token=WIKI_PARENT_TOKEN,
        object_token=WIKI_PARENT_OBJECT_TOKEN,
        object_type="docx",
        title="测试用例",
        parent_node_token=None,
    )


def _xlsx(spec: SheetDeliverySpec) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "空白页"
    for row_index, row in enumerate(spec.remote_values(), start=1):
        worksheet.row_dimensions[row_index].height = 18
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row_index, column_index, value)
            cell.font = Font(
                bold=False,
                italic=False,
                size=10,
                strike=False,
                color="FF000000",
            )
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=False
            )
    for column in ("A", "B"):
        worksheet.column_dimensions[column].width = (100 - 5) / 7
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _target() -> ProtectedTarget:
    return ProtectedTarget(
        source_locator=LOCATOR,
        spreadsheet_token=SHEET_TOKEN,
        workbook_title="目标工作簿",
        workbook_url=f"https://example.feishu.cn/sheets/{SHEET_TOKEN}",
        worksheet_selector=SHEET_ID,
        sheet_id=SHEET_ID,
        sheet_title="空白页",
        sheet_index=0,
        initial_revision="7",
        initial_sheet_count=1,
        initial_state_hash="sha256:initial",
    )


def _recovery_record(spec: SheetDeliverySpec) -> OperationRecord:
    now = datetime(2026, 8, 25, 0, 0, tzinfo=UTC)
    return OperationRecord(
        operation_ref="wop_0123456789abcdef0123456789abcdef",
        task_ref="task-recovery",
        profile_ref=PROFILE_REF,
        placement_mode=PlacementMode.CREATE_NEW_SHEET,
        spec_hash=spec.content_hash,
        preview_hash="sha256:preview",
        delivery_hash=None,
        target=_target(),
        state=OperationState.RECOVERY_REQUIRED,
        last_completed_step=WriteStep.GRID_EXTENDED,
        ambiguous=True,
        remote_revision=None,
        diagnostic_code="values_write_contract_unknown",
        created_at=now,
        expires_at=now + timedelta(minutes=10),
        updated_at=now,
    )


def _xlsx_with_wrong_first_value(spec: SheetDeliverySpec) -> bytes:
    workbook = load_workbook(BytesIO(_xlsx(spec)))
    worksheet = workbook["空白页"]
    worksheet["A1"] = "错误值"
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _xlsx_with_column_widths(
    spec: SheetDeliverySpec,
    widths: tuple[float, ...],
) -> bytes:
    workbook = load_workbook(BytesIO(_xlsx(spec)))
    worksheet = workbook["空白页"]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[chr(ord("A") + index - 1)].width = width
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _revision_xlsx(spec: SheetDeliverySpec) -> bytes:
    workbook = load_workbook(BytesIO(_xlsx(spec)))
    worksheet = workbook["空白页"]
    worksheet.row_dimensions[3].height = 18
    worksheet.column_dimensions["C"].width = (100 - 5) / 7
    for coordinate in ("C1", "C2", "A3", "B3", "C3"):
        worksheet[coordinate].alignment = Alignment(
            horizontal="left",
            vertical="top",
        )
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_adoption_preview_uses_exact_selector_and_proves_content_blank() -> None:
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        assert request.headers["authorization"] == (
            "Bearer write-access-token-must-not-render"
        )
        if request.url.path.endswith(f"/spreadsheets/{SHEET_TOKEN}"):
            return httpx.Response(200, json=_metadata())
        if request.url.path.endswith("/sheets/query"):
            return httpx.Response(
                200,
                json={"code": 0, "data": {"sheets": [_sheet(), {
                    **_sheet(),
                    "sheet_id": "sheet-two",
                    "title": "另一个表",
                    "index": 1,
                }]}},
            )
        assert request.url.path.endswith("/values_batch_get")
        assert request.url.params["ranges"] == f"{SHEET_ID}!A1:B2"
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "revision": 7,
                    "valueRanges": [
                        {"range": f"{SHEET_ID}!A1:B2", "values": []}
                    ],
                },
            },
        )

    lease = FakeLeaseClient()
    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(lease_client=lease, http_client=http)
    result = asyncio.run(
        gateway.preview(
            locator=LOCATOR,
            task_ref="task-preview",
            profile_ref=PROFILE_REF,
            placement_mode=PlacementMode.ADOPT_BLANK_SHEET,
            requested_sheet_title=None,
            spec=_spec(),
        )
    )
    asyncio.run(http.aclose())

    assert result.target.sheet_id == SHEET_ID
    assert result.target.sheet_title == "空白页"
    assert result.target.initial_revision == "7"
    assert result.warnings == ("non_atomic_blank_sheet_adoption",)
    assert len(requests) == 3
    assert lease.calls[0][2] == (
        SHEETS_EXPORT_VERIFY_CAPABILITY,
        SHEETS_MANAGED_WRITE_CAPABILITY,
        SHEETS_READ_CAPABILITY,
    )


def test_create_preview_requires_typed_write_scope_for_boolean_values() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith(f"/spreadsheets/{SHEET_TOKEN}"):
            return httpx.Response(200, json=_metadata())
        assert request.url.path.endswith("/sheets/query")
        return httpx.Response(
            200, json={"code": 0, "data": {"sheets": [_sheet()]}}
        )

    lease = FakeLeaseClient()
    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(lease_client=lease, http_client=http)
    asyncio.run(
        gateway.preview(
            locator=f"https://example.feishu.cn/sheets/{SHEET_TOKEN}",
            task_ref="task-typed-preview",
            profile_ref=PROFILE_REF,
            placement_mode=PlacementMode.CREATE_NEW_SHEET,
            requested_sheet_title="布尔用例",
            spec=_boolean_spec(),
        )
    )
    asyncio.run(http.aclose())

    assert lease.calls[0][2] == (
        SHEETS_EXPORT_VERIFY_CAPABILITY,
        SHEETS_MANAGED_WRITE_CAPABILITY,
        SHEETS_READ_CAPABILITY,
        SHEETS_TYPED_VALUES_WRITE_CAPABILITY,
    )


def test_adoption_preview_rejects_nonblank_content() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith(f"/spreadsheets/{SHEET_TOKEN}"):
            return httpx.Response(200, json=_metadata())
        if request.url.path.endswith("/sheets/query"):
            return httpx.Response(
                200, json={"code": 0, "data": {"sheets": [_sheet()]}}
            )
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "valueRanges": [
                        {
                            "range": f"{SHEET_ID}!A1:B2",
                            "values": [[{"type": "url", "text": "链接"}]],
                        }
                    ]
                },
            },
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(CapabilityError) as error:
        asyncio.run(
            gateway.preview(
                locator=LOCATOR,
                task_ref="task-preview",
                profile_ref=PROFILE_REF,
                placement_mode=PlacementMode.ADOPT_BLANK_SHEET,
                requested_sheet_title=None,
                spec=_spec(),
            )
        )
    asyncio.run(http.aclose())

    assert error.value.code is CapabilityErrorCode.PRECONDITION_FAILED


def test_create_preview_uses_workbook_only_and_requires_a_unique_title() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith(f"/spreadsheets/{SHEET_TOKEN}"):
            return httpx.Response(200, json=_metadata())
        assert request.url.path.endswith("/sheets/query")
        return httpx.Response(
            200, json={"code": 0, "data": {"sheets": [_sheet()]}}
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    result = asyncio.run(
        gateway.preview(
            locator=LOCATOR,
            task_ref="task-create-preview",
            profile_ref=PROFILE_REF,
            placement_mode=PlacementMode.CREATE_NEW_SHEET,
            requested_sheet_title="新建用例",
            spec=_spec(),
        )
    )

    assert result.target.sheet_id is None
    assert result.target.requested_sheet_title == "新建用例"
    assert result.warnings == ("worksheet_selector_ignored_for_create",)

    with pytest.raises(CapabilityError) as collision:
        asyncio.run(
            gateway.preview(
                locator=LOCATOR,
                task_ref="task-create-collision",
                profile_ref=PROFILE_REF,
                placement_mode=PlacementMode.CREATE_NEW_SHEET,
                requested_sheet_title="空白页",
                spec=_spec(),
            )
        )
    asyncio.run(http.aclose())

    assert collision.value.code is CapabilityErrorCode.WRITE_CONFLICT


def test_new_workbook_preview_is_zero_write_and_binds_complete_wiki_children() -> None:
    requests: list[httpx.Request] = []

    def handler(request: httpx.Request) -> httpx.Response:
        requests.append(request)
        assert request.method == "GET"
        if request.url.path.endswith("/spaces/get_node"):
            return httpx.Response(
                200,
                json={"code": 0, "data": {"node": _wiki_parent()}},
            )
        assert request.url.path.endswith(f"/spaces/{WIKI_SPACE_ID}/nodes")
        assert request.url.params["parent_node_token"] == WIKI_PARENT_TOKEN
        if "page_token" not in request.url.params:
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "items": [],
                        "has_more": True,
                        "page_token": "nextPageToken123",
                    },
                },
            )
        assert request.url.params["page_token"] == "nextPageToken123"
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "items": [
                        _wiki_node(
                            node_token="wikcnExistingChild123",
                            object_token="doccnExistingChild123",
                            object_type="docx",
                            title="已有说明",
                        )
                    ],
                    "has_more": False,
                },
            },
        )

    lease = FakeLeaseClient()
    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(lease_client=lease, http_client=http)
    result = asyncio.run(
        gateway.preview(
            locator=WIKI_PARENT_LOCATOR,
            task_ref="task-new-workbook-preview",
            profile_ref=PROFILE_REF,
            placement_mode=PlacementMode.CREATE_NEW_WORKBOOK,
            requested_sheet_title=None,
            requested_workbook_title="正式用例-20260828",
            spec=_spec(),
        )
    )
    asyncio.run(http.aclose())

    assert result.target.spreadsheet_token is None
    assert result.target.requested_workbook_title == "正式用例-20260828"
    assert result.target.wiki_space_id == WIKI_SPACE_ID
    assert result.target.parent_wiki_node_token == WIKI_PARENT_TOKEN
    assert result.target.initial_child_count == 1
    assert result.target.sheet_id is None
    assert len(requests) == 3
    assert lease.calls[0][2] == (
        SHEETS_EXPORT_VERIFY_CAPABILITY,
        SHEETS_MANAGED_WRITE_CAPABILITY,
        SHEETS_READ_CAPABILITY,
        WIKI_NODE_READ_CAPABILITY,
        WIKI_CHILD_LIST_CAPABILITY,
        WIKI_NODE_CREATE_CAPABILITY,
    )


def test_new_workbook_preview_stops_on_sibling_title_collision() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("/spaces/get_node"):
            return httpx.Response(
                200,
                json={"code": 0, "data": {"node": _wiki_parent()}},
            )
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "items": [
                        _wiki_node(
                            node_token="wikcnConflictChild123",
                            object_token="shtcnConflictChild123",
                            object_type="sheet",
                            title="正式用例-20260828",
                        )
                    ],
                    "has_more": False,
                },
            },
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(CapabilityError) as error:
        asyncio.run(
            gateway.preview(
                locator=WIKI_PARENT_LOCATOR,
                task_ref="task-new-workbook-conflict",
                profile_ref=PROFILE_REF,
                placement_mode=PlacementMode.CREATE_NEW_WORKBOOK,
                requested_sheet_title=None,
                requested_workbook_title="正式用例-20260828",
                spec=_spec(),
            )
        )
    asyncio.run(http.aclose())

    assert error.value.code is CapabilityErrorCode.WRITE_CONFLICT


def test_new_workbook_creation_checkpoints_file_then_adopts_default_sheet() -> None:
    created_body: dict[str, object] = {}

    def preview_handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("/spaces/get_node"):
            return httpx.Response(
                200,
                json={"code": 0, "data": {"node": _wiki_parent()}},
            )
        return httpx.Response(
            200,
            json={"code": 0, "data": {"items": [], "has_more": False}},
        )

    preview_http = httpx.AsyncClient(transport=httpx.MockTransport(preview_handler))
    preview_gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=preview_http
    )
    preview = asyncio.run(
        preview_gateway.preview(
            locator=WIKI_PARENT_LOCATOR,
            task_ref="task-new-workbook-create",
            profile_ref=PROFILE_REF,
            placement_mode=PlacementMode.CREATE_NEW_WORKBOOK,
            requested_sheet_title=None,
            requested_workbook_title="正式用例-20260828",
            spec=_spec(),
        )
    )
    asyncio.run(preview_http.aclose())

    def apply_handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("/spaces/get_node"):
            return httpx.Response(
                200,
                json={"code": 0, "data": {"node": _wiki_parent()}},
            )
        if request.url.path.endswith(f"/spaces/{WIKI_SPACE_ID}/nodes"):
            if request.method == "GET":
                return httpx.Response(
                    200,
                    json={"code": 0, "data": {"items": [], "has_more": False}},
                )
            created_body.update(json.loads(request.content))
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "node": _wiki_node(
                            node_token=CREATED_WIKI_NODE_TOKEN,
                            object_token=CREATED_SHEET_TOKEN,
                            object_type="sheet",
                            title="正式用例-20260828",
                        )
                    },
                },
            )
        if request.url.path.endswith(f"/spreadsheets/{CREATED_SHEET_TOKEN}"):
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "spreadsheet": {
                            "title": "正式用例-20260828",
                            "token": CREATED_SHEET_TOKEN,
                            "url": f"https://example.feishu.cn/sheets/{CREATED_SHEET_TOKEN}",
                        }
                    },
                },
            )
        if request.url.path.endswith("/sheets/query"):
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "sheets": [
                            {
                                **_sheet(),
                                "sheet_id": "created-default-sheet",
                                "title": "Sheet1",
                            }
                        ]
                    },
                },
            )
        assert request.url.path.endswith("/values_batch_get")
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "revision": 3,
                    "valueRanges": [
                        {
                            "range": "created-default-sheet!A1:B2",
                            "values": [],
                        }
                    ],
                },
            },
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(apply_handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    created = asyncio.run(
        gateway._recheck_and_create_workbook(preview.target, {"Authorization": "Bearer test"})
    )
    registered, revision = asyncio.run(
        gateway._register_created_workbook_target(
            created,
            {"Authorization": "Bearer test"},
        )
    )
    asyncio.run(http.aclose())

    assert created_body == {
        "obj_type": "sheet",
        "parent_node_token": WIKI_PARENT_TOKEN,
        "node_type": "origin",
        "title": "正式用例-20260828",
    }
    assert created.spreadsheet_token == CREATED_SHEET_TOKEN
    assert created.created_wiki_node_token == CREATED_WIKI_NODE_TOKEN
    assert registered.sheet_id == "created-default-sheet"
    assert registered.sheet_title == "Sheet1"
    assert registered.workbook_url == (
        f"https://example.feishu.cn/sheets/{CREATED_SHEET_TOKEN}"
    )
    assert revision == "3"


def test_ambiguous_new_workbook_create_reconciles_exact_new_child() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("/spaces/get_node"):
            return httpx.Response(
                200,
                json={"code": 0, "data": {"node": _wiki_parent()}},
            )
        items = []
        if handler.list_calls:
            items.append(
                _wiki_node(
                    node_token=CREATED_WIKI_NODE_TOKEN,
                    object_token=CREATED_SHEET_TOKEN,
                    object_type="sheet",
                    title="正式用例-20260828",
                )
            )
        handler.list_calls += 1
        return httpx.Response(
            200,
            json={"code": 0, "data": {"items": items, "has_more": False}},
        )

    handler.list_calls = 0  # type: ignore[attr-defined]
    lease = FakeLeaseClient()
    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(lease_client=lease, http_client=http)
    preview = asyncio.run(
        gateway.preview(
            locator=WIKI_PARENT_LOCATOR,
            task_ref="task-new-workbook-recovery",
            profile_ref=PROFILE_REF,
            placement_mode=PlacementMode.CREATE_NEW_WORKBOOK,
            requested_sheet_title=None,
            requested_workbook_title="正式用例-20260828",
            spec=_spec(),
        )
    )
    now = datetime(2026, 8, 28, 0, 0, tzinfo=UTC)
    record = OperationRecord(
        operation_ref="wop_abcdefabcdefabcdefabcdefabcdefab",
        task_ref="task-new-workbook-recovery",
        profile_ref=PROFILE_REF,
        placement_mode=PlacementMode.CREATE_NEW_WORKBOOK,
        spec_hash=_spec().content_hash,
        preview_hash="sha256:preview",
        delivery_hash=None,
        target=preview.target,
        state=OperationState.RECOVERY_REQUIRED,
        last_completed_step=WriteStep.NONE,
        ambiguous=True,
        remote_revision=None,
        diagnostic_code="workbook_create_transport_unknown",
        created_at=now,
        expires_at=now + timedelta(minutes=10),
        updated_at=now,
    )
    progress = asyncio.run(gateway.reconcile_progress(record=record, spec=_spec()))
    asyncio.run(http.aclose())

    assert progress is not None
    assert progress.completed_step is WriteStep.WORKBOOK_CREATED
    assert progress.target is not None
    assert progress.target.spreadsheet_token == CREATED_SHEET_TOKEN
    assert progress.target.created_wiki_node_token == CREATED_WIKI_NODE_TOKEN
    assert progress.warnings == (
        "ambiguous_workbook_create_reconciled_by_child_diff",
    )
    assert lease.calls[-1][2] == (
        WIKI_NODE_READ_CAPABILITY,
        WIKI_CHILD_LIST_CAPABILITY,
    )


def test_create_sheet_registers_only_the_stable_returned_sheet_identity() -> None:
    seen_body: dict[str, object] = {}

    def handler(request: httpx.Request) -> httpx.Response:
        seen_body.update(json.loads(request.content))
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "revision": 8,
                    "replies": [
                        {
                            "addSheet": {
                                "properties": {
                                    "sheetId": "created-sheet",
                                    "title": "新建用例",
                                    "index": 1,
                                }
                            }
                        }
                    ],
                },
            },
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    initial = ProtectedTarget(
        source_locator=LOCATOR,
        spreadsheet_token=SHEET_TOKEN,
        workbook_title="目标工作簿",
        workbook_url=f"https://example.feishu.cn/sheets/{SHEET_TOKEN}",
        worksheet_selector=SHEET_ID,
        requested_sheet_title="新建用例",
        initial_sheet_count=1,
        initial_state_hash="sha256:initial",
    )
    target, revision = asyncio.run(
        gateway._add_sheet(
            initial,
            "新建用例",
            1,
            {"Authorization": "Bearer test"},
        )
    )
    asyncio.run(http.aclose())

    assert seen_body == {
        "requests": [
            {"addSheet": {"properties": {"title": "新建用例", "index": 1}}}
        ]
    }
    assert target.sheet_id == "created-sheet"
    assert target.sheet_title == "新建用例"
    assert target.sheet_index == 1
    assert revision == "8"


@pytest.mark.parametrize(
    "payload",
    (
        {"code": 0, "msg": "success"},
        {"code": "0", "msg": "success", "data": None},
    ),
)
def test_values_write_accepts_explicit_success_without_data_object(
    payload: dict[str, object],
) -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        assert request.method == "PUT"
        assert request.url.path.endswith(f"/spreadsheets/{SHEET_TOKEN}/values")
        return httpx.Response(200, json=payload)

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    revision = asyncio.run(
        gateway._write_values(
            _target(),
            _spec(),
            {"Authorization": "Bearer test"},
        )
    )
    asyncio.run(http.aclose())

    assert revision is None


def test_values_write_rejects_missing_explicit_platform_success() -> None:
    def handler(_: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json={"msg": "success"})

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(RemoteMutationFailure) as error:
        asyncio.run(
            gateway._write_values(
                _target(),
                _spec(),
                {"Authorization": "Bearer test"},
            )
        )
    asyncio.run(http.aclose())

    assert error.value.diagnostic_code == "values_write_contract_unknown"
    assert error.value.ambiguous is True


def test_values_write_preserves_explicit_platform_rejection() -> None:
    def handler(_: httpx.Request) -> httpx.Response:
        return httpx.Response(
            200,
            json={"code": 90204, "msg": "valueRange is wrong", "data": {}},
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(RemoteMutationFailure) as error:
        asyncio.run(
            gateway._write_values(
                _target(),
                _spec(),
                {"Authorization": "Bearer test"},
            )
        )
    asyncio.run(http.aclose())

    assert error.value.diagnostic_code == (
        "values_write_provider_contract_error:90204"
    )
    assert error.value.ambiguous is False


def test_values_write_with_boolean_uses_typed_cell_tool() -> None:
    spec = _boolean_spec()

    def handler(request: httpx.Request) -> httpx.Response:
        assert request.method == "POST"
        assert request.url.path.endswith(
            f"/sheet_ai/v2/spreadsheets/{SHEET_TOKEN}/tools/invoke_write"
        )
        body = json.loads(request.content)
        assert body["tool_name"] == "set_cell_range"
        tool_input = json.loads(body["input"])
        assert tool_input == {
            "excel_id": SHEET_TOKEN,
            "sheet_id": SHEET_ID,
            "range": "A1:B2",
            "cells": [
                [{"value": "布尔"}, {}],
                [{"value": True}, {"value": 2}],
            ],
        }
        return httpx.Response(
            200,
            json={"code": 0, "msg": "success", "data": {"output": "{}"}},
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    revision = asyncio.run(
        gateway._write_values(
            _target(),
            spec,
            {"Authorization": "Bearer test"},
        )
    )
    asyncio.run(http.aclose())

    assert revision is None


def test_typed_values_write_preserves_missing_scope_platform_rejection() -> None:
    def handler(_: httpx.Request) -> httpx.Response:
        return httpx.Response(
            400,
            json={
                "code": 99991679,
                "msg": "sensitive missing-scope provider message",
                "error": {},
            },
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(RemoteMutationFailure) as error:
        asyncio.run(
            gateway._write_values(
                _target(),
                _boolean_spec(),
                {"Authorization": "Bearer test"},
            )
        )
    asyncio.run(http.aclose())

    assert error.value.diagnostic_code == "values_write_auth_required:99991679"
    assert error.value.ambiguous is False
    assert "sensitive missing-scope" not in str(error.value)


def test_reconcile_progress_proves_ambiguous_values_write_by_api_readback() -> None:
    spec = _spec()

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("/sheets/query"):
            return httpx.Response(
                200, json={"code": 0, "data": {"sheets": [_sheet()]}}
            )
        assert request.url.path.endswith("/values_batch_get")
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "revision": 19,
                    "valueRanges": [
                        {
                            "range": f"{SHEET_ID}!A1:B2",
                            "revision": 19,
                            "values": spec.remote_values(),
                        }
                    ],
                },
            },
        )

    lease = FakeLeaseClient()
    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(lease_client=lease, http_client=http)
    progress = asyncio.run(
        gateway.reconcile_progress(record=_recovery_record(spec), spec=spec)
    )
    asyncio.run(http.aclose())

    assert progress is not None
    assert progress.completed_step is WriteStep.VALUES_WRITTEN
    assert progress.remote_revision == "19"
    assert progress.warnings == (
        "ambiguous_values_write_reconciled_by_api_readback",
    )
    assert lease.calls[0][2] == (SHEETS_READ_CAPABILITY,)


@pytest.mark.parametrize(
    ("remote_values", "expected_step"),
    (
        ([], WriteStep.GRID_EXTENDED),
        ([["partial"]], None),
    ),
)
def test_reconcile_progress_only_resumes_when_values_are_fully_blank(
    remote_values: list[list[object]],
    expected_step: WriteStep | None,
) -> None:
    spec = _spec()

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path.endswith("/sheets/query"):
            return httpx.Response(
                200, json={"code": 0, "data": {"sheets": [_sheet()]}}
            )
        assert request.url.path.endswith("/values_batch_get")
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "revision": 20,
                    "valueRanges": [
                        {
                            "range": f"{SHEET_ID}!A1:B2",
                            "revision": 20,
                            "values": remote_values,
                        }
                    ],
                },
            },
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    progress = asyncio.run(
        gateway.reconcile_progress(record=_recovery_record(spec), spec=spec)
    )
    asyncio.run(http.aclose())

    if expected_step is None:
        assert progress is None
    else:
        assert progress is not None
        assert progress.completed_step is expected_step
        assert progress.remote_revision == "20"
        assert progress.warnings == (
            "ambiguous_values_write_proved_not_applied_by_api_readback",
        )


def test_export_reports_safe_create_permission_diagnostic() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        assert request.url.path == "/open-apis/drive/v1/export_tasks"
        return httpx.Response(
            403,
            json={"code": 1069902, "msg": "no permission"},
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(RemoteMutationFailure) as error:
        asyncio.run(
            gateway._verify_export(
                _target(),
                _spec(),
                {"Authorization": "Bearer test"},
            )
        )
    asyncio.run(http.aclose())

    assert error.value.diagnostic_code == (
        "xlsx_export_create_permission_denied:1069902"
    )
    assert error.value.verification_incomplete is True
    assert error.value.ambiguous is False


def test_export_reports_terminal_job_status_diagnostic() -> None:
    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == "/open-apis/drive/v1/export_tasks":
            return httpx.Response(
                200,
                json={"code": 0, "data": {"ticket": "ticket123"}},
            )
        assert request.url.path.endswith("/export_tasks/ticket123")
        return httpx.Response(
            200,
            json={
                "code": 0,
                "data": {
                    "result": {
                        "job_status": 110,
                        "job_error_msg": "must not be persisted",
                    }
                },
            },
        )

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(RemoteMutationFailure) as error:
        asyncio.run(
            gateway._verify_export(
                _target(),
                _spec(),
                {"Authorization": "Bearer test"},
            )
        )
    asyncio.run(http.aclose())

    assert error.value.diagnostic_code == "xlsx_export_job_permission_denied"
    assert "must not be persisted" not in error.value.diagnostic_code


def test_export_reports_first_safe_xlsx_mismatch() -> None:
    spec = _spec()
    exported = _xlsx_with_wrong_first_value(spec)

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == "/open-apis/drive/v1/export_tasks":
            return httpx.Response(
                200,
                json={"code": 0, "data": {"ticket": "ticket123"}},
            )
        if request.url.path.endswith("/export_tasks/ticket123"):
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "result": {
                            "job_status": 0,
                            "file_token": "filetoken123",
                            "file_size": len(exported),
                        }
                    },
                },
            )
        assert request.url.path.endswith("/file/filetoken123/download")
        return httpx.Response(200, content=exported)

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(RemoteMutationFailure) as error:
        asyncio.run(
            gateway._verify_export(
                _target(),
                spec,
                {"Authorization": "Bearer test"},
            )
        )
    asyncio.run(http.aclose())

    assert error.value.diagnostic_code == "xlsx_verify_cell_value:0:0"


def test_export_compacts_multiple_dimension_mismatches_safely() -> None:
    spec = _spec()
    exported = _xlsx_with_column_widths(spec, (19.0, 12.0))

    def handler(request: httpx.Request) -> httpx.Response:
        if request.url.path == "/open-apis/drive/v1/export_tasks":
            return httpx.Response(
                200,
                json={"code": 0, "data": {"ticket": "ticket123"}},
            )
        if request.url.path.endswith("/export_tasks/ticket123"):
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "result": {
                            "job_status": 0,
                            "file_token": "filetoken123",
                            "file_size": len(exported),
                        }
                    },
                },
            )
        assert request.url.path.endswith("/file/filetoken123/download")
        return httpx.Response(200, content=exported)

    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=FakeLeaseClient(), http_client=http
    )
    with pytest.raises(RemoteMutationFailure) as error:
        asyncio.run(
            gateway._verify_export(
                _target(),
                spec,
                {"Authorization": "Bearer test"},
            )
        )
    asyncio.run(http.aclose())

    assert error.value.diagnostic_code == (
        "xlsx_verify_dimensions:c0-19000-138-100_c1-12000-89-100"
    )


def test_execute_serializes_mutations_then_api_and_xlsx_readback() -> None:
    spec = _spec()
    exported = _xlsx(spec)
    mutations: list[tuple[str, dict[str, object]]] = []
    query_count = 0

    def handler(request: httpx.Request) -> httpx.Response:
        nonlocal query_count
        if request.url.path.endswith("/sheets/query"):
            query_count += 1
            return httpx.Response(
                200, json={"code": 0, "data": {"sheets": [_sheet()]}}
            )
        if request.url.path.endswith("/values_batch_get"):
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "revision": 18,
                        "valueRanges": [
                            {
                                "range": f"{SHEET_ID}!A1:B2",
                                "revision": 18,
                                "values": spec.remote_values(),
                            }
                        ],
                    },
                },
            )
        if request.url.path == "/open-apis/drive/v1/export_tasks":
            mutations.append(("export", json.loads(request.content)))
            return httpx.Response(200, json={"code": 0, "data": {"ticket": "ticket123"}})
        if request.url.path.endswith("/export_tasks/ticket123"):
            assert request.url.params["token"] == SHEET_TOKEN
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "result": {
                            "job_status": 0,
                            "file_token": "filetoken123",
                            "file_size": len(exported),
                        }
                    },
                },
            )
        if request.url.path.endswith("/file/filetoken123/download"):
            return httpx.Response(200, content=exported)

        body = json.loads(request.content)
        mutations.append((request.url.path, body))
        return httpx.Response(200, json={"code": 0, "data": {"revision": 17}})

    lease = FakeLeaseClient()
    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=lease,
        http_client=http,
        export_poll_interval_seconds=0,
    )
    now = datetime(2026, 8, 25, 0, 0, tzinfo=UTC)
    target = ProtectedTarget(
        source_locator=LOCATOR,
        spreadsheet_token=SHEET_TOKEN,
        workbook_title="目标工作簿",
        workbook_url=f"https://example.feishu.cn/sheets/{SHEET_TOKEN}",
        worksheet_selector=SHEET_ID,
        sheet_id=SHEET_ID,
        sheet_title="空白页",
        sheet_index=0,
        initial_revision="7",
        initial_sheet_count=1,
        initial_state_hash="sha256:initial",
    )
    record = OperationRecord(
        operation_ref="wop_0123456789abcdef0123456789abcdef",
        task_ref="task-apply",
        profile_ref=PROFILE_REF,
        placement_mode=PlacementMode.ADOPT_BLANK_SHEET,
        spec_hash=spec.content_hash,
        preview_hash="sha256:preview",
        delivery_hash=None,
        target=target,
        state=OperationState.EXECUTING,
        last_completed_step=WriteStep.TARGET_REGISTERED,
        ambiguous=False,
        remote_revision="7",
        diagnostic_code=None,
        created_at=now,
        expires_at=now + timedelta(minutes=10),
        updated_at=now,
    )
    steps: list[WriteStep] = []
    proof = asyncio.run(
        gateway.execute(
            record=record,
            spec=spec,
            on_workbook=lambda *_: pytest.fail("workbook was already registered"),
            on_target=lambda *_: pytest.fail("target was already registered"),
            on_step=lambda step, _: steps.append(step),
        )
    )
    asyncio.run(http.aclose())

    assert proof.api_hash.startswith("sha256:")
    assert proof.export_hash.startswith("sha256:")
    assert proof.remote_revision == "18"
    assert query_count == 2
    assert steps == [
        WriteStep.GRID_EXTENDED,
        WriteStep.VALUES_WRITTEN,
        WriteStep.STYLES_CLEARED,
        WriteStep.BASE_STYLE_WRITTEN,
        WriteStep.STYLE_RANGES_WRITTEN,
        WriteStep.DIMENSIONS_WRITTEN,
        WriteStep.FREEZE_WRITTEN,
        WriteStep.MERGES_WRITTEN,
        WriteStep.API_VERIFIED,
        WriteStep.EXPORT_VERIFIED,
    ]
    assert mutations[0][1] == {
        "valueRange": {
            "range": f"{SHEET_ID}!A1:B2",
            "values": [["标题", "值"], [1, 2]],
        }
    }
    assert mutations[1][1]["data"][0]["style"] == {"clean": True}
    assert mutations[2][1]["data"][0]["style"]["font"]["fontSize"] == "10pt/1.5"
    assert mutations[-1] == (
        "export",
        {"file_extension": "xlsx", "token": SHEET_TOKEN, "type": "sheet"},
    )


def test_execute_revision_cleans_retired_area_and_rebuilds_state() -> None:
    base = SheetDeliverySpec.model_validate(
        {
            **_spec().model_dump(mode="json"),
            "row_count": 3,
            "column_count": 3,
            "values": [
                ["旧标题", "", "退役"],
                [1, 2, 3],
                [4, 5, 6],
            ],
            "merges": [
                {
                    "row_start": 0,
                    "row_end": 1,
                    "column_start": 0,
                    "column_end": 2,
                }
            ],
        }
    )
    next_spec = _spec()
    diff = build_revision_diff(base, next_spec)
    exported = _revision_xlsx(next_spec)
    mutations: list[tuple[str, dict[str, object]]] = []
    base_merge_present = True

    def sheet_payload() -> dict[str, object]:
        payload = _sheet()
        payload["grid_properties"] = {
            "row_count": 3,
            "column_count": 3,
            "frozen_row_count": 0,
            "frozen_column_count": 0,
        }
        payload["merges"] = (
            [
                    {
                        "start_row_index": 0,
                        "end_row_index": 0,
                        "start_column_index": 0,
                        "end_column_index": 1,
                    }
            ]
            if base_merge_present
            else []
        )
        return payload

    def handler(request: httpx.Request) -> httpx.Response:
        nonlocal base_merge_present
        if request.url.path.endswith(f"/spreadsheets/{SHEET_TOKEN}"):
            return httpx.Response(200, json=_metadata())
        if request.url.path.endswith("/sheets/query"):
            return httpx.Response(
                200,
                json={"code": 0, "data": {"sheets": [sheet_payload()]}},
            )
        if request.url.path.endswith("/values_batch_get"):
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "revision": 30,
                        "valueRanges": [
                            {
                                "range": f"{SHEET_ID}!A1:C3",
                                "revision": 30,
                                "values": next_spec.remote_values(),
                            }
                        ],
                    },
                },
            )
        if request.url.path == "/open-apis/drive/v1/export_tasks":
            mutations.append(("export", json.loads(request.content)))
            return httpx.Response(
                200, json={"code": 0, "data": {"ticket": "ticket123"}}
            )
        if request.url.path.endswith("/export_tasks/ticket123"):
            return httpx.Response(
                200,
                json={
                    "code": 0,
                    "data": {
                        "result": {
                            "job_status": 0,
                            "file_token": "filetoken123",
                            "file_size": len(exported),
                        }
                    },
                },
            )
        if request.url.path.endswith("/file/filetoken123/download"):
            return httpx.Response(200, content=exported)

        body = json.loads(request.content)
        mutations.append((request.url.path, body))
        if request.url.path.endswith("/unmerge_cells"):
            base_merge_present = False
        return httpx.Response(200, json={"code": 0, "data": {"revision": 29}})

    lease = FakeLeaseClient()
    http = httpx.AsyncClient(transport=httpx.MockTransport(handler))
    gateway = FeishuManagedSheetsGateway(
        lease_client=lease,
        http_client=http,
        export_poll_interval_seconds=0,
    )
    now = datetime(2026, 8, 25, 0, 0, tzinfo=UTC)
    registration = ManagedSheetRegistration(
        registration_ref="managed_0123456789abcdef0123456789abcdef",
        initial_operation_ref="wop_0123456789abcdef0123456789abcdef",
        profile_ref=PROFILE_REF,
        target=_target(),
        current_version=1,
        spec_hash=base.content_hash,
        spec_summary=base.summary(),
        delivery_hash="sha256:" + "1" * 64,
        remote_revision="20",
        state=OperationState.DELIVERED,
        created_at=now,
        updated_at=now,
    )
    record = RevisionRecord(
        operation_ref="rev_0123456789abcdef0123456789abcdef",
        registration_ref=registration.registration_ref,
        task_ref="task-revision",
        base_spec_hash=base.content_hash,
        next_spec_hash=next_spec.content_hash,
        next_spec_summary=next_spec.summary(),
        preview_hash="sha256:" + "2" * 64,
        diff_summary=diff.model_dump(mode="json"),
        candidate_version=2,
        state=RevisionState.EXECUTING,
        last_completed_step=RevisionStep.REVISION_RESERVED,
        ambiguous=False,
        remote_revision="20",
        diagnostic_code=None,
        base_api_hash="sha256:" + "3" * 64,
        base_export_hash="sha256:" + "4" * 64,
        delivery_hash=None,
        created_at=now,
        expires_at=now + timedelta(minutes=10),
        updated_at=now,
    )
    steps: list[RevisionStep] = []
    proof = asyncio.run(
        gateway.execute_revision(
            record=record,
            registration=registration,
            base_spec=base,
            next_spec=next_spec,
            diff=diff,
            on_step=lambda step, _: steps.append(step),
        )
    )
    recovery_record = replace(
        record,
        state=RevisionState.VERIFICATION_INCOMPLETE,
        last_completed_step=RevisionStep.API_VERIFIED,
        diagnostic_code="xlsx_verify_retired_style:2:0:vertical_alignment",
    )
    reconciled = asyncio.run(
        gateway.reconcile_revision_final(
            record=recovery_record,
            registration=registration,
            base_spec=base,
            next_spec=next_spec,
            diff=diff,
        )
    )
    asyncio.run(http.aclose())

    assert proof.remote_revision == "30"
    assert reconciled is not None
    assert reconciled.remote_revision == "30"
    assert reconciled.warnings == (
        "retired_styles_normalized_during_reconciliation",
    )
    assert SHEETS_MANAGED_WRITE_CAPABILITY in lease.calls[-1][2]
    assert steps == list(tuple(RevisionStep)[2:-1])
    unmerge = next(
        body for path, body in mutations if path.endswith("/unmerge_cells")
    )
    assert unmerge == {"range": f"{SHEET_ID}!A1:B1"}
    retired_writes = [
        body
        for path, body in mutations
        if path.endswith(f"/spreadsheets/{SHEET_TOKEN}/values")
    ][1:]
    assert [item["valueRange"]["range"] for item in retired_writes] == [
        f"{SHEET_ID}!A3:C3",
        f"{SHEET_ID}!C1:C2",
    ]
    dimension_bodies = [
        body for path, body in mutations if path.endswith("/dimension_range")
    ]
    assert dimension_bodies[-2]["dimensionProperties"]["fixedSize"] == 24
    assert dimension_bodies[-1]["dimensionProperties"]["fixedSize"] == 100
    retired_style_body = next(
        body
        for path, body in mutations
        if path.endswith("/styles_batch_update")
        and any(
            item["style"].get("borderType") == "NO_BORDER"
            for item in body["data"]
        )
    )
    assert retired_style_body == {
        "data": [
            {
                "ranges": [f"{SHEET_ID}!A3:C3"],
                "style": {
                    "font": {"clean": True},
                    "textDecoration": 0,
                    "formatter": "",
                    "hAlign": 0,
                    "vAlign": 0,
                    "borderType": "NO_BORDER",
                    "clean": False,
                },
            },
            {
                "ranges": [f"{SHEET_ID}!C1:C2"],
                "style": {
                    "font": {"clean": True},
                    "textDecoration": 0,
                    "formatter": "",
                    "hAlign": 0,
                    "vAlign": 0,
                    "borderType": "NO_BORDER",
                    "clean": False,
                },
            },
        ]
    }
    assert mutations[-1][0] == "export"
